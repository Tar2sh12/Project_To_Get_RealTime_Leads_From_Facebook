from selenium import webdriver
from selenium.webdriver.common.by import By
import time
import schedule
import openpyxl
# Define your function
def my_function(driver):
    try:
        data_dict = []
        driver.refresh()
        time.sleep(2)
        firstElement = driver.find_elements(By.CSS_SELECTOR, '.x1i10hfl.x1qjc9v5.xjqpnuy.xa49m3k.xqeqjp1.x2hbi6w.x972fbf.xcfux6l.x1qhh985.xm0m39n.x9f619.x1ypdohk.xdl72j9.xe8uvvx.xdj266r.x11i5rnm.xat24cr.x1mh8g0r.x2lwn1j.xeuugli.x16tdsg8.xggy1nq.x1ja2u2z.x1n2onr6.x1q0g3np.xxymvpz.x1ejq31n.xd10rxx.x1sy0etr.x17r0tee.x87ps6o.x1t137rt.xlh3980.xvmahel.x1hl2dhg.x1lku1pv.x78zum5.x1iyjqo2.xs83m0k.x1lcm9me.x1yr5g0i.xrt01vj.x10y3i5r.xo1l8bm.xbsr9hj.x1v911su.x1y1aw1k.xwib8y2.x1swvt13.x1pi30zi')
        for x in firstElement:
            x.click()
            time.sleep(2)
            dict = {}
            name = driver.find_element(By.XPATH, '/html/body/div[1]/div/div[1]/div/div[2]/div/div/div[1]/span/div/div/div[1]/div[1]/div/div/div/div/div/div/div/div/div/div[2]/div/div/div/div[1]/div[2]/div/div/div/div/div[1]/div/div/div[1]/div')
            email = driver.find_element(By.XPATH,'/html/body/div[1]/div/div[1]/div/div[2]/div/div/div[1]/span/div/div/div[1]/div[1]/div/div/div/div/div/div/div/div/div/div[2]/div/div/div/div[1]/div[2]/div/div/div/div/div[2]/div[1]/div[3]/div[1]/div[1]/div[3]')
            phone = driver.find_element(By.XPATH,'/html/body/div[1]/div/div[1]/div/div[2]/div/div/div[1]/span/div/div/div[1]/div[1]/div/div/div/div/div/div/div/div/div/div[2]/div/div/div/div[1]/div[2]/div/div/div/div/div[2]/div[1]/div[3]/div[1]/div[1]/div[2]')
            date = driver.find_element(By.XPATH,'/html/body/div[1]/div/div[1]/div/div[2]/div/div/div[1]/span/div/div/div[1]/div[1]/div/div/div/div/div/div/div/div/div/div[2]/div/div/div/div[1]/div[2]/div/div/div/div/div[2]/div[1]/div[2]')
            dict['name']=name.text
            dict['email'] = email.text
            dict['phone'] = phone.text
            # data_dict['date']= date.text.split()[3] + " " + date.split()[4]
            # print(f"name : {name.text} \nemail : {email.text} \nphone : {phone.text} \ndate : {date.text} \n ")
            data_dict.append(dict)
            time.sleep(2)
        file_path = r'C:\Users\admin\PycharmProjects\leadsAPI\leads.xlsx'
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        # Write headers
        headers = list(data_dict[0].keys())
        for col_index, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col_index).value = header

        # Write data rows
        for row_index, entry in enumerate(data_dict, start=2):
            for col_index, key in enumerate(headers, start=1):
                sheet.cell(row=row_index, column=col_index).value = entry[key]
        print(data_dict)
        workbook.save(file_path)
    except Exception as e:
        print(f"An error occurred: {str(e)}")

# Initialize WebDriver
driver = webdriver.Chrome()
driver.get('https://business.facebook.com/business/loginpage/?next=https%3A%2F%2Fbusiness.facebook.com%2Flatest%2Fleads_center%3Fasset_id%3D404684342708751%26nav_ref%3Dbiz_unified_f3_login_page_to_mbs&login_options%5B0%5D=FB&login_options%5B1%5D=IG&login_options%5B2%5D=SSO&config_ref=biz_login_tool_flavor_mbs')
time.sleep(3)

# Login
loginElement = driver.find_element(By.XPATH,'/html/body/div/div[1]/div/div[1]/div/div/div/div[2]/div/div/div/div[2]/div/div/div/div')
loginElement.click()
time.sleep(3)
usernameXPATH= '/html/body/div[1]/div[2]/div[1]/div/div[2]/div[2]/form/div[2]/div[1]/input'
passXPATH='/html/body/div[1]/div[2]/div[1]/div/div[2]/div[2]/form/div[2]/div[2]/div/div/input'
login_buttonXPATH='/html/body/div[1]/div[2]/div[1]/div/div[2]/div[2]/form/div[2]/div[3]/button'
driver.find_element(By.XPATH,usernameXPATH).send_keys('')#your email
driver.find_element(By.XPATH,passXPATH).send_keys('')#your pass
driver.find_element(By.XPATH,login_buttonXPATH).click()
time.sleep(5)

# Navigate to Leads Center
driver.get('https://business.facebook.com/latest/leads_center?asset_id=404684342708751')
time.sleep(5)

# Schedule the function to run every minute
schedule.every(1).minutes.do(my_function, driver)

# Loop to continuously check and run scheduled tasks
while True:
    schedule.run_pending()
    time.sleep(1)  # Adjust sleep time as needed


